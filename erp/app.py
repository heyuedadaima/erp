from flask import Flask, jsonify, request,send_file
from sqlalchemy import create_engine, text
from flask_cors import CORS
import os
from datetime import datetime
import pdfplumber
from openpyxl import load_workbook
import traceback
import pandas as pd
import sys
import io
import pymysql


JSON_AS_ASCII = False
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')



app = Flask(__name__)
CORS(app)

# 配置数据库连接信息
DATABASE_CONFIG = {
    'user': 'root',  # 替换为你的数据库用户名
    'password': '123456',  # 替换为你的数据库密码
    'host':'mysql',  # 数据库地址
    'port': 3306,  # 数据库端口
    'database': 'erp'  # 替换为你的数据库名称
}

# 数据库连接字符串
DB_URL = f"mysql+pymysql://{DATABASE_CONFIG['user']}:{DATABASE_CONFIG['password']}@" \
         f"{DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}/{DATABASE_CONFIG['database']}"

# 创建数据库引擎
engine = create_engine(DB_URL)

@app.before_request
def initialize_database():
    try:
        message = create_database_if_not_exists()  # 每次请求都检查数据库
        print(message)
    except Exception as e:
        print(str(e))

def create_database_if_not_exists():
    connection = None
    try:
        # 获取当前脚本所在目录
        current_directory = os.path.dirname(__file__)

        # 构建 SQL 文件路径
        sql_file_path = os.path.join(current_directory, 'erp3.sql')

        # 连接到 MySQL 服务器（没有指定数据库名称）
        connection = pymysql.connect(
            host=DATABASE_CONFIG['host'],
            user=DATABASE_CONFIG['user'],
            password=DATABASE_CONFIG['password'],
            port=DATABASE_CONFIG['port']
        )

        with connection.cursor() as cursor:
            # 检查是否存在 `erp` 数据库
            cursor.execute("SHOW DATABASES LIKE 'erp'")
            result = cursor.fetchone()

            if result is None:
                # 如果不存在，创建数据库
                cursor.execute("CREATE DATABASE erp")
                print("Database 'erp' created.")

                # 切换到 `erp` 数据库
                cursor.execute("USE erp")

                # 读取 SQL 脚本文件
                with open(sql_file_path, 'r') as file:
                    sql_script = file.read()

                    # 将 SQL 脚本按分号拆分为单独的语句
                    sql_statements = sql_script.split(';')

                    # 逐条执行 SQL 语句
                    for statement in sql_statements:
                        if statement.strip():  # 跳过空语句
                            try:
                                cursor.execute(statement)
                            except pymysql.MySQLError as e:
                                print(f"Error executing SQL statement: {statement}")
                                raise

                connection.commit()
                return "Database 'erp' created and SQL script executed successfully."
            else:
                # 如果数据库存在，检查是否为空表
                cursor.execute("USE erp")
                cursor.execute("SHOW TABLES")
                tables = cursor.fetchall()

                if not tables:
                    # 如果数据库为空表，执行 SQL 脚本
                    print("Database 'erp' exists but is empty. Executing SQL script...")

                    # 读取 SQL 脚本文件
                    with open(sql_file_path, 'r') as file:
                        sql_script = file.read()

                        # 将 SQL 脚本按分号拆分为单独的语句
                        sql_statements = sql_script.split(';')

                        # 逐条执行 SQL 语句
                        for statement in sql_statements:
                            if statement.strip():  # 跳过空语句
                                try:
                                    cursor.execute(statement)
                                except pymysql.MySQLError as e:
                                    print(f"Error executing SQL statement: {statement}")
                                    raise

                    connection.commit()
                    return "Database 'erp' is empty. SQL script executed successfully."
                else:
                    return "Database 'erp' already exists and is not empty."

    except pymysql.MySQLError as e:
        raise Exception(f"MySQL Error: {str(e)}")
    except FileNotFoundError as fnf_error:
        raise Exception(f"File Error: {str(fnf_error)}")
    except Exception as e:
        raise Exception(f"Error creating database or executing script: {str(e)}")
    finally:
        if connection:
            connection.close()



@app.route('/products', methods=['GET'])
def get_products():
    try:
        with engine.connect() as connection:
            query = text("SELECT name, address FROM product")
            result = connection.execute(query)
            products = [{'name': row[0], 'address': row[1]} for row in result.fetchall()]

        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/products', methods=['DELETE'])
def delete_products():
    try:
        # 获取请求体中的要删除的产品信息（name 或 address）
        product_data = request.json.get('products', [])

        if not product_data:
            return jsonify({'error': '没有提供要删除的产品信息'}), 400

        # 遍历要删除的每个产品，生成 DELETE SQL 查询
        with engine.connect() as connection:
            for product in product_data:
                name = product.get('name')
                address = product.get('address')

                # 删除文件（如果文件存在）
                if os.path.exists(address):
                    os.remove(address)


                # 构建删除 SQL 查询
                query = text("DELETE FROM product WHERE name = :name AND address = :address")


                # 执行删除操作
                result = connection.execute(query, {"name": name, "address": address})

                connection.commit()
        return jsonify({'message': '删除成功'})

    except Exception as e:
        print(f"删除失败: {str(e)}")  # 打印错误信息
        return jsonify({'error': str(e)}), 500



UPLOAD_FOLDER = 'uploads'  # 上传文件保存的目录
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER



@app.route('/upload', methods=['POST'])
def upload_file():
    # 获取传递的 name 和 文件
    name = request.form.get('name')  # 获取传递的 name
    file = request.files.get('file')  # 获取上传的文件
    if not name or not file:
        return {'message': '缺少必要的参数（name 或 file）'}, 400
        # 获取文件的原始名称和扩展名
    filename, file_extension = os.path.splitext(file.filename)

        # 重命名文件，使用传递的 name 字段值作为文件名
    new_filename = f"{name}{file_extension}"

        # 保存文件
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
    file.save(filepath)

    # 将文件信息插入到数据库中
    with engine.connect() as connection:
        query = text("INSERT INTO product (name, address) VALUES (:name, :address)")
        connection.execute(query, {"name": name, "address": filepath})
        connection.commit()

    return {
        'message': '文件上传成功并已保存到数据库',
        'name': name,
        'file_path': filepath
    }, 200


@app.route('/orders', methods=['GET'])
def get_orders():
    try:
        # 获取查询参数
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        date = request.args.get('date')  # 获取传递的日期

        offset = (page - 1) * page_size

        # 基本查询，按制单日期从大到小排序
        query = "SELECT * FROM `order`"
        if date:
            query += " WHERE 制单日期 = :date"
        query += " ORDER BY 制单日期 DESC"

        # 计算总记录数
        count_query = "SELECT COUNT(*) FROM `order`"
        if date:
            count_query += " WHERE 制单日期 = :date"

        with engine.connect() as connection:
            # 获取总记录数
            total_count_result = connection.execute(text(count_query), {"date": date})
            total_count = total_count_result.scalar()

            # 获取分页数据
            query += " LIMIT :offset, :page_size"
            result = connection.execute(text(query), {"date": date, "offset": offset, "page_size": page_size})

            orders = [dict(zip(result.keys(), row)) for row in result.fetchall()]

        return jsonify({"orders": orders, "total_count": total_count})
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500




@app.route('/orders', methods=['DELETE'])
def delete_orders():
    try:
        order_data = request.json.get('products', [])
        if not order_data:
            return jsonify({'error': '没有提供要删除的订单信息'}), 400

        with engine.connect() as connection:
            for order in order_data:
                order_id = order.get('序列')  # 使用 '序列' 字段作为删除条件

                # 打印 SQL 语句和参数，确保没有问题
                print(f"准备删除: 序列={order_id}")

                # 删除数据库记录
                query = text("DELETE FROM `order` WHERE 序列 = :order_id")  # 使用 '序列' 字段进行删除
                result = connection.execute(query, {"order_id": order_id})
                connection.commit()



        return jsonify({'message': '删除成功'})
    except Exception as e:
      
        return jsonify({'error': str(e)}), 500

@app.route('/orders', methods=['POST'])
def add_order():
    try:
        # 获取请求体中的订单数据
        order_data = request.json

        # 获取制单日期，如果没有则使用当前日期
        created_at = order_data.get('制单日期', datetime.now().strftime('%Y-%m-%d'))

        # 将订单数据插入到数据库
        with engine.connect() as connection:
            query = text("""
                INSERT INTO `order` (模号, 件号, 品名, 规格, 数量, 客户, 制单日期)
                VALUES (:模号, :件号, :品名, :规格, :数量, :客户, :制单日期)
            """)
            connection.execute(query, {
                '模号': order_data['模号'],
                '件号': order_data['件号'],
                '品名': order_data['品名'],
                '规格': order_data['规格'],
                '数量': order_data['数量'],
                '客户': order_data['客户'],
                '制单日期': created_at,
            })
            connection.commit()

        return jsonify({'message': '订单添加成功'}), 201

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500





# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'pdf', 'xls', 'xlsx'}

# 检查文件类型是否允许上传
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 确保数据文件夹存在
DATA_FOLDER = 'print'
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)

uploaded_data = []

@app.route('/upload1', methods=['POST'])
def upload_file1():
    if 'file' not in request.files:
        return jsonify({'message': '没有文件 part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': '没有选择文件'}), 400

    if not allowed_file(file.filename):
        return jsonify({'message': '不支持的文件类型'}), 400

    file = request.files['file']
    print(f"Uploaded file name: {file.filename}")
    # 获取文件的原始名称和扩展名
    filename = file.filename

    # 保存到 data 文件夹
    file_path = os.path.join(DATA_FOLDER, filename)
    file.save(file_path)

    try:
        # 解析文件并提取表格内容
        file_extension = filename.rsplit('.', 1)[1].lower()
        columns = []
        global uploaded_data  # 使用全局变量
        if file_extension in ['xls', 'xlsx']:
            columns, uploaded_data = process_excel(file_path)
        elif file_extension == 'pdf':
            columns, uploaded_data = process_pdf(file_path)

        print(columns)
        print(uploaded_data)

        # 返回列名映射和数据给前端
        return jsonify({'columnsMap': columns, 'data': uploaded_data}), 200
    finally:
        # 确保处理完文件后删除文件
        if os.path.exists(file_path):
            os.remove(file_path)


# 处理 Excel 文件
def process_excel(file_path):
    try:
        # 读取 Excel 文件
        df = pd.read_excel(file_path, header=0, engine='openpyxl')



        # 将 Excel 数据转换为字典形式
        columns = df.columns.tolist()
        data = df.to_dict(orient='records')

        # 可以在此进行额外的数据格式检查和转换，比如确保所有数字列是数字类型
        for row in data:
            # 如果某列必须是数字，可以在这里强制转换
            try:
                row['数量'] = float(row.get('数量', 0))
            except ValueError:
                row['数量'] = 0  # 如果转换失败，填充为 0

        return columns, data

    except Exception as e:
        print(f"处理 Excel 文件时发生错误：{e}")
        raise

# 处理 PDF 文件
def process_pdf(file_path):
    columns = []
    data = []

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                # 假设第一行是表头
                if not columns:
                    columns = table[0]  # 第一页的第一行作为表头
                for row in table[1:]:
                    # 检查是否是空行
                    if not any(cell.strip() for cell in row if cell):
                        break  # 如果遇到空行，停止读取后续行
                    # 收集行数据
                    data.append(dict(zip(columns, row)))

    return columns, data



@app.route('/process-uploaded-file', methods=['POST'])
def process_uploaded_file():
    # 获取列映射
    columns_map = request.json.get('columnsMap', [])
    client=request.json.get('client')
    print(1)
    print(client)
    if not columns_map:
        return jsonify({'error': '没有提供列映射'}), 400

    global uploaded_data  # 使用全局变量
    if not uploaded_data:
        return jsonify({'error': '没有文件数据'}), 400

    # 进行数据映射和插入
    try:
        with engine.connect() as connection:
            for row in uploaded_data:  # `uploaded_data` 是上传的文件解析结果（每行数据是一个字典）
                order_data = {}

                # 遍历 columns_map，将文件数据映射到数据库列
                for column_mapping in columns_map:
                    # 获取数据库列名（databaseColumn）和文件列名（mappedColumn）
                    database_column = column_mapping.get('databaseColumn')
                    mapped_column = column_mapping.get('mappedColumn')

                    # 获取对应的文件列数据
                    if mapped_column in row:
                        order_data[database_column] = row[mapped_column]
                    else:
                        order_data[database_column] = None  # 如果没有对应的数据，设置为 None

                # 自动填入制单日期（当天日期）
                order_data['制单日期'] = datetime.now().strftime('%Y-%m-%d')
                order_data['客户']=client

                # 执行插入数据库操作
                query = text(""" 
                    INSERT INTO `order` (模号, 件号, 品名, 规格, 数量, 客户, 制单日期) 
                    VALUES (:模号, :件号, :品名, :规格, :数量, :客户, :制单日期) 
                """)
                connection.execute(query, order_data)

            connection.commit()

        return jsonify({'message': '数据处理成功并已保存到数据库'}), 200
    except Exception as e:
        return jsonify({'error': f'数据插入失败: {str(e)}'}), 500







UPLOAD_FOLDER = './data'


@app.route('/print-labels', methods=['POST'])
def print_labels():
    try:
        # 获取请求数据
        request_data = request.json
        data = request_data.get('rows', [])  # 获取 rows 列表

        # 验证数据
        if not data:
            return jsonify({"message": "未提供 rows 数据"}), 400

        # 定义插入位置，每行插入到的列（右侧一列）
        positions = [
            ('B3', 'B4', 'B5', 'B7', 'B8'),  # 第一行
            ('E3', 'E4', 'E5', 'E7', 'E8'),  # 第二行
            ('B11', 'B12', 'B13', 'B15', 'B16'),  # 第三行
            ('E11', 'E12', 'E13', 'E15', 'E16'),  # 第四行
            ('B19', 'B20', 'B21', 'B23', 'B24'),  # 第五行
            ('E19', 'E20', 'E21', 'E23', 'E24'),
            ('B27', 'B28', 'B29', 'B31', 'B32'),
            ('E27', 'E28', 'E29', 'E31', 'E32'),
            ('B35', 'B36', 'B37', 'B39', 'B40'),
            ('E35', 'E36', 'E37', 'E39', 'E40'),
        ]

        # 分批处理数据，每次插入 10 行
        max_rows = len(data)
        chunk_size = 10
        num_files = (max_rows // chunk_size) + (1 if max_rows % chunk_size > 0 else 0)
        file_paths = []

        # 分批生成文件
        for file_index in range(num_files):
            # 每个文件的行数据
            start_idx = file_index * chunk_size
            end_idx = min((file_index + 1) * chunk_size, max_rows)
            rows_to_insert = data[start_idx:end_idx]

            # 创建新的 Excel 文件
            template_path = os.path.join('./uploads', '格式.xlsx')
            wb = load_workbook(template_path)
            sheet = wb.active

            # 处理每行数据，插入到 Excel 文件中
            for row_index, row in enumerate(rows_to_insert):
                # 获取每一行的数据
                customer = row.get('客户', '')
                product = row.get('品名', '')
                spec = row.get('规格', '')
                quantity = row.get('数量', '')
                model = row.get('模号', '')

                # 获取该行对应的列位置
                current_positions = positions[row_index % len(positions)]  # 循环使用插入位置

                # 将数据插入到对应的右侧列
                sheet[current_positions[0]] = customer
                sheet[current_positions[1]] = product
                sheet[current_positions[2]] = spec
                sheet[current_positions[3]] = quantity
                sheet[current_positions[4]] = model

            # 保存文件到 data 文件夹中
            file_name = f"打印标签_{file_index + 1}.xlsx"
            file_path = os.path.join(UPLOAD_FOLDER, file_name)
            wb.save(file_path)
            file_paths.append(file_path)



        # 返回成功消息，包含生成的文件路径
        return jsonify({"message": "打印成功！", "file_paths": file_paths})

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            "message": f"服务器内部错误: {str(e)}",
            "debug": traceback.format_exc()
        }), 500


UPLOAD_FOLDER = './data'

@app.route('/download-printed-labels', methods=['GET'])
def download_printed_labels():
    try:
        # 获取data文件夹中所有Excel文件
        excel_files = [f for f in os.listdir(UPLOAD_FOLDER) if f.endswith('.xlsx')]

        # 如果没有找到文件，返回错误信息
        if not excel_files:
            return jsonify({"message": "没有找到打印标签文件"}), 404

        # 返回第一个Excel文件供下载
        file_path = os.path.join(UPLOAD_FOLDER, excel_files[0])  # 返回第一个生成的Excel文件
        return send_file(file_path, as_attachment=True, download_name='printed_labels.xlsx')

    except Exception as e:
        return jsonify({"message": f"服务器错误: {str(e)}"}), 500



@app.route('/endpoint', methods=['POST'])
def handle_order_data():
    try:
        # 获取请求数据
        data = request.get_json()
        print("Received data:", data)  # 打印接收到的数据
        print(data)

        select_data = data.get('selectData')  # 获取订单数据
        outbound_format = data.get('outboundFormat')  # 获取出库格式
        print(select_data)
        print(outbound_format)
        # 检查必填字段
        if not select_data or not outbound_format:
            return jsonify({'error': 'selectData 和 outboundFormat 为必填项'}), 400

        print("Received selectData:", select_data)
        print("Received outboundFormat:", outbound_format)

        # 根据出库格式选择模板文件
        template_folder = './uploads'
        output_folder = './endprint'
        template_file = f"{outbound_format}.xlsx"  # 模板文件名，如 '福耀出货单.xlsx'
        template_path = os.path.join(template_folder, template_file)

        # 检查模板文件是否存在
        if not os.path.exists(template_path):
            return jsonify({'error': f'模板文件未找到: {template_path}'}), 404

        # 加载模板文件
        workbook = load_workbook(template_path)
        sheet = workbook.active  # 获取活动的工作表

        # 填充数据
        if outbound_format == "锦运出货单":
            fill_jinyun_outbound(sheet, select_data)
        elif outbound_format == "福耀出货单":
            fill_fuyao_outbound(sheet, select_data)
        elif outbound_format == "三捷出货单":
            fill_sanjie_outbound(sheet, select_data)
        else:
            return jsonify({'error': '未知的出库格式'}), 400

        # 确保输出目录存在
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 设置输出文件名
        output_file = "出库单.xlsx"
        output_path = os.path.join(output_folder, output_file)

        # 保存生成的 Excel 文件
        workbook.save(output_path)


        # 返回生成的文件，供前端下载
        return send_file(
            output_path,
            as_attachment=True,  # 表示以附件形式下载
            download_name=output_file,  # 下载时的文件名
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        # 捕获异常并返回详细错误信息
        print("Error occurred:", str(e))
        return jsonify({'error': f"发生错误: {str(e)}"}), 500


# 针对锦运出货单的填充规则
def fill_jinyun_outbound(sheet, select_data):
    # 填充数据
    row_start = 7  # 从第7行开始填充数据
    row_end = row_start + len(select_data) - 1  # 根据数据长度确定结束行，最多填充10行
    if row_end > 17:  # 确保填充的行数不超过10行
        row_end = 17

    for index, item in enumerate(select_data):
        row = row_start + index  # 每个item填充一行数据
        # 填充对应单元格
        sheet[f'A{row}'] = item.get('模号', '')
        sheet[f'B{row}'] = item.get('品名', '')
        sheet[f'C{row}'] = item.get('规格', '')
        sheet[f'D{row}'] = '支'  # 单位固定为'支'
        sheet[f'E{row}'] = item.get('数量', 0)

# 针对福耀出货单的填充规则
def fill_fuyao_outbound(sheet, select_data):
    row_start = 5  # 福耀出货单从第5行开始填充
    row_end = row_start + len(select_data) - 1
    if row_end > 14:  # 确保填充的行数不超过10行
        row_end = 14

    for index, item in enumerate(select_data):
        row = row_start + index  # 每个item填充一行数据
        # 填充对应单元格
        sheet[f'A{row}'] = row - 4  # 序号，从1开始自动增长
        sheet[f'B{row}'] = item.get('件号')
        sheet[f'C{row}'] = item.get('品名', '')
        sheet[f'D{row}'] = item.get('规格', '')
        sheet[f'E{row}'] = ' '
        sheet[f'G{row}'] = item.get('数量', 0)
        sheet[f'H{row}'] = item.get('模号', '')

# 针对三捷出货单的填充规则
def fill_sanjie_outbound(sheet, select_data):
    # 填充数据
    row_start = 13  # 三捷出货单从第13行开始填充
    row_end = row_start + len(select_data) - 1
    if row_end > 25:  # 确保填充的行数不超过13行
        row_end = 25

    for index, item in enumerate(select_data):
        row = row_start + index  # 每个item填充一行数据
        # 填充对应单元格
        sheet[f'D{row}'] = item.get('模号', '')  # 模号填充到D列
        sheet[f'E{row}'] = item.get('品名', '')  # 品名填充到E列
        sheet[f'G{row}'] = item.get('规格', '')  # 规格填充到G列
        sheet[f'I{row}'] = item.get('数量', 0)  # 数量填充到I列






if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)